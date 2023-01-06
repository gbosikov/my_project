import openpyxl
from pathlib import Path

xlsx_file = Path('TEST.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active

row_list = []
clear_val_list = []
for row in sheet.iter_rows(1, sheet.max_row):
    val_in_cell = row[0].value.replace(';', '').replace('/', '')
    if len(val_in_cell) > 9:
        clear_right_val = val_in_cell[9:]
        clear_left_val = val_in_cell[:9]
        clear_val_list.append(clear_left_val)
        clear_val_list.append(clear_right_val)
    else:
        clear_val_list.append(val_in_cell)


for val in clear_val_list:
    if val != str(0):
        print(val)






